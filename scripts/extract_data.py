import csv
import re
import sys
import unicodedata
import zipfile
from datetime import date, datetime
from pathlib import Path
from xml.etree import ElementTree as ET

try:
    from openpyxl import load_workbook
except ImportError:
    load_workbook = None


def _xlsx_cell_value(cell, shared_strings):
    cell_type = cell.get("t")
    if cell_type == "inlineStr":
        text = cell.find("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}is")
        if text is None:
            return ""
        return "".join(node.text or "" for node in text.iter())
    if cell_type == "s":
        index = int(cell.findtext("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v", "0"))
        return shared_strings[index] if 0 <= index < len(shared_strings) else ""
    if cell_type == "b":
        return cell.findtext("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v", "0") == "1"
    return cell.findtext("{http://schemas.openxmlformats.org/spreadsheetml/2006/main}v", "")


def _read_xlsx_fallback(path):
    class FakeWorkbook:
        def __init__(self, sheets):
            self.worksheets = sheets

    class FakeWorksheet:
        def __init__(self, title, rows):
            self.title = title
            self._rows = rows

        def iter_rows(self, min_row=1, values_only=True):
            for index, row in enumerate(self._rows, start=1):
                if index < min_row:
                    continue
                if values_only:
                    yield tuple(row)
                else:
                    yield row

    ns_main = "http://schemas.openxmlformats.org/spreadsheetml/2006/main"
    ns_rel = "http://schemas.openxmlformats.org/officeDocument/2006/relationships"
    ns_pkg = "http://schemas.openxmlformats.org/package/2006/relationships"

    with zipfile.ZipFile(path) as xlsx:
        shared_strings = []
        if "xl/sharedStrings.xml" in xlsx.namelist():
            root = ET.fromstring(xlsx.read("xl/sharedStrings.xml"))
            for node in root.findall(f"{{{ns_main}}}si"):
                shared_strings.append("".join(t.text or "" for t in node.iter() if t.tag == f"{{{ns_main}}}t"))

        workbook = ET.fromstring(xlsx.read("xl/workbook.xml"))
        rels = ET.fromstring(xlsx.read("xl/_rels/workbook.xml.rels"))
        rel_map = {r.attrib["Id"]: r.attrib["Target"] for r in rels.findall(f"{{{ns_pkg}}}Relationship")}

        worksheets = []
        for sheet in workbook.findall(f"{{{ns_main}}}sheets/{{{ns_main}}}sheet"):
            rel_id = sheet.attrib.get(f"{{{ns_rel}}}id")
            target = rel_map.get(rel_id, "")
            if not target:
                continue
            sheet_path = target if target.startswith("/") else f"xl/{target}"
            sheet_root = ET.fromstring(xlsx.read(sheet_path))
            rows = []
            for row in sheet_root.findall(f".//{{{ns_main}}}row"):
                row_values = []
                for cell in row.findall(f"{{{ns_main}}}c"):
                    cell_ref = cell.attrib.get("r", "")
                    if cell_ref and cell_ref[0].isalpha():
                        row_values.append(_xlsx_cell_value(cell, shared_strings))
                    else:
                        row_values.append(_xlsx_cell_value(cell, shared_strings))
                if row_values:
                    rows.append(row_values)
            worksheets.append(FakeWorksheet(sheet.attrib.get("name", "Sheet"), rows))

    return FakeWorkbook(worksheets)


SOURCE = Path(sys.argv[1])
OUTPUT = Path(sys.argv[2])

def text(value):
    if value is None:
        return ""
    if isinstance(value, (datetime, date)):
        return value.strftime("%Y-%m-%d")
    return re.sub(r"\s+", " ", str(value).replace("\xa0", " ")).strip()

def clean_status(value):
    value = text(value)
    aliases = {
        "no inicido": "No iniciado",
        "no iniciado": "No iniciado",
        "en curso": "En curso",
        "terminado": "Terminado",
        "rechazado": "Rechazado",
    }
    return aliases.get(value.strip().lower(), value.strip() or "Sin estado")

def split_people(value):
    value = text(value)
    if not value or value.lower() in {"todos", "por definir"}:
        return [value] if value else []
    value = re.sub(r"\s+y\s+", ",", value, flags=re.I)
    people = [p.strip() for p in value.split(",") if p.strip()]
    people = [{"Martin": "Martín"}.get(person, person) for person in people]
    return list(dict.fromkeys(people))

if load_workbook is not None:
    wb = load_workbook(SOURCE, data_only=True, read_only=True)
else:
    wb = _read_xlsx_fallback(SOURCE)

projects = []
people_rows = []

for sheet in wb.worksheets:
    for row in sheet.iter_rows(min_row=2, values_only=True):
        if not row[0]:
            continue
        values = list(row[:7]) + [None] * max(0, 7 - len(row))
        project_id = f"P{len(projects)+1:03d}"
        people = split_people(values[1])
        record = {
            "id": project_id,
            "categoria": sheet.title,
            "proyecto": text(values[0]),
            "responsables": text(values[1]),
            "lista_responsables": "|".join(people),
            "fecha_inicio": text(values[2]),
            "fecha_fin": text(values[3]),
            "estado": clean_status(values[4]),
            "observaciones": text(values[5]),
            "por_hacer": text(values[6]),
        }
        projects.append(record)
        people_rows.extend({"id": project_id, "responsable": person} for person in people)

OUTPUT.mkdir(parents=True, exist_ok=True)
with (OUTPUT / "proyectos.csv").open("w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=projects[0].keys())
    writer.writeheader(); writer.writerows(projects)
with (OUTPUT / "responsables.csv").open("w", encoding="utf-8-sig", newline="") as f:
    writer = csv.DictWriter(f, fieldnames=["id", "responsable"])
    writer.writeheader(); writer.writerows(people_rows)

print(f"{len(projects)} proyectos y {len(people_rows)} asignaciones exportadas")
